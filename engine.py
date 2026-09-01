"""
Motor genérico de ampliación de materiales a centros.

La idea: un "tipo" (Modelos, Repuestos, ...) se define solo con un JSON en
config/tipos/*.json. Este motor no conoce reglas de negocio de ningún tipo
en particular; todo sale de la config + la tabla de referencia (Excel).

Para agregar un tipo nuevo (ej. Repuestos) más adelante:
  1. Crear config/tipos/repuestos.json con sus propias columnas de input,
     su propio archivo de referencia y su propia key de matching.
  2. Dejar el archivo de referencia en data/reference/.
  3. Listo — aparece automático en el selector de la web.
"""

import json
from pathlib import Path

import openpyxl
import pandas as pd

import correlativo

BASE_DIR = Path(__file__).parent
TIPOS_DIR = BASE_DIR / "config" / "tipos"


class TipoNoEncontradoError(Exception):
    pass


class InputInvalidoError(Exception):
    pass


def listar_tipos() -> list[dict]:
    tipos = []
    for f in sorted(TIPOS_DIR.glob("*.json")):
        with open(f, encoding="utf-8") as fh:
            cfg = json.load(fh)
        tipos.append({"id": cfg["id"], "nombre": cfg["nombre"], "descripcion": cfg.get("descripcion", "")})
    return tipos


def cargar_config(tipo_id: str) -> dict:
    path = TIPOS_DIR / f"{tipo_id}.json"
    if not path.exists():
        raise TipoNoEncontradoError(f"No existe configuración para el tipo '{tipo_id}'.")
    with open(path, encoding="utf-8") as fh:
        return json.load(fh)


def generar_plantilla_vacia(tipo_id: str, filas_vacias: int = 200) -> "openpyxl.Workbook":
    """
    Arma un Excel vacío con la hoja y columnas de input del tipo (solo
    encabezados), para que alguien sin el Excel original tenga de dónde
    partir. Las columnas de texto (config "text_columns") quedan con
    formato de celda texto ('@') en las filas vacías, para evitar el
    problema de siempre (ceros iniciales / notación científica) sin tener
    que explicarle al usuario que las formatee él mismo.
    """
    cfg = cargar_config(tipo_id)
    columnas = cfg["input_columns"]
    text_columns = set(cfg.get("text_columns", []))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = cfg["input_sheet"]
    ws.append(columnas)

    for row in range(2, 2 + filas_vacias):
        for col_idx, nombre_col in enumerate(columnas, start=1):
            if nombre_col in text_columns:
                ws.cell(row=row, column=col_idx).number_format = "@"

    for col_idx, nombre_col in enumerate(columnas, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(14, len(nombre_col) + 2)

    diccionario_cfg = cfg.get("diccionario_referencia")
    if diccionario_cfg:
        _agregar_hoja_diccionario(wb, diccionario_cfg)

    return wb


def _agregar_hoja_diccionario(wb: "openpyxl.Workbook", diccionario_cfg: dict) -> None:
    """Copia tal cual una hoja de referencia (ej. marcas/jerarquía) como hoja
    extra de la plantilla descargable, para que el usuario tenga a mano los
    códigos válidos sin tener que abrir otro Excel aparte."""
    origen_path = BASE_DIR / diccionario_cfg["reference_file"]
    if not origen_path.exists():
        return
    wb_origen = openpyxl.load_workbook(origen_path, data_only=True)
    ws_origen = wb_origen[diccionario_cfg["sheet"]]

    ws_destino = wb.create_sheet(title=diccionario_cfg.get("titulo_hoja", "DICCIONARIO"))
    for fila in ws_origen.iter_rows(values_only=True):
        ws_destino.append(fila)


def _leer_input(file_storage, cfg: dict) -> pd.DataFrame:
    # Preparar dtype dict y converters para especificar columnas de texto
    text_columns = cfg.get("text_columns", [])
    dtype_dict = {col: str for col in text_columns}
    converters_dict = {col: str for col in text_columns}
    
    try:
        df = pd.read_excel(
            file_storage, 
            sheet_name=cfg["input_sheet"], 
            dtype=dtype_dict,
            converters=converters_dict
        )
    except ValueError as e:
        raise InputInvalidoError(
            f"No se encontró la hoja '{cfg['input_sheet']}' en el Excel subido."
        ) from e

    df = df.dropna(how="all")
    df.columns = [str(c).strip() for c in df.columns]

    columnas_esperadas = cfg["input_columns"]
    faltantes = [c for c in columnas_esperadas if c not in df.columns]
    if faltantes:
        raise InputInvalidoError(
            "El Excel de input no tiene las columnas esperadas: " + ", ".join(faltantes)
        )

    # Solo nos quedamos con las columnas definidas por la plantilla, en su orden,
    # y descartamos cualquier fila que venga completamente vacía en la key de match.
    df = df[columnas_esperadas].copy()

    # Cada tipo puede llamar distinto a su columna de filial (Modelos: "FILIAL
    # CODIGO", Repuestos: "FILIAL"...). Hacia adentro del sistema siempre se
    # trabaja con "FILIAL CODIGO", en mayúsculas (Vc00, vC00, VC00... se tratan
    # igual), para no tener que enseñarle "FILIAL" a app.py/salida_sap.py.
    columna_filial = cfg.get("columna_filial", "FILIAL CODIGO")
    if columna_filial in df.columns:
        df[columna_filial] = df[columna_filial].astype(str).str.strip().str.upper()
        if columna_filial != "FILIAL CODIGO":
            df = df.rename(columns={columna_filial: "FILIAL CODIGO"})

    # Convertir columnas especificadas a texto para evitar notación científica y pérdida de ceros
    for col in text_columns:
        if col in df.columns:
            df[col] = df[col].astype(str)
    
    df = df.dropna(subset=[cfg["key_input"]])
    if df.empty:
        raise InputInvalidoError("El Excel no tiene ninguna fila de datos para procesar.")
    
    return df


def _cargar_referencia(ref_file: str, ref_sheet) -> pd.DataFrame:
    ref_path = BASE_DIR / ref_file
    if not ref_path.exists():
        raise InputInvalidoError(f"Falta el archivo de referencia: {ref_path}")
    return pd.read_excel(ref_path, sheet_name=ref_sheet or 0)


def _cargar_disponibilidad(cfg: dict) -> dict[tuple[str, str], float] | None:
    """
    Matriz Centro x Fabricante confirmada por negocio (1 = habilitado,
    0.5 = por confirmar pero se trata como habilitado, 0 = bloqueado: ese
    fabricante no tiene CEBE para ese centro y NO puede generarse ahí,
    aunque la tabla de referencia fabricante->centro/cebe tenga una fila
    para esa combinación (puede ser un error de carga en esa tabla).

    Layout fijo de la hoja (ver data/reference/Centros_UN_RP.xlsx, hoja
    "Unidades"): fila 1 = códigos de fabricante desde la columna 3 en
    adelante, filas 3+ = un centro por fila (columna 1 = CENTRO).

    Devuelve None si el tipo no configuró "disponibilidad" (nadie valida
    nada, comportamiento anterior sin cambios).
    """
    disp_cfg = cfg.get("disponibilidad")
    if not disp_cfg:
        return None

    path = BASE_DIR / disp_cfg["reference_file"]
    if not path.exists():
        raise InputInvalidoError(f"Falta el archivo de disponibilidad: {path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[disp_cfg["sheet"]]

    fabricantes = [ws.cell(row=1, column=c).value for c in range(3, ws.max_column + 1)]
    lookup: dict[tuple[str, str], float] = {}
    for r in range(3, ws.max_row + 1):
        centro = ws.cell(row=r, column=1).value
        if not centro:
            continue
        for i, fab in enumerate(fabricantes):
            if not fab:
                continue
            valor = ws.cell(row=r, column=3 + i).value
            if valor is not None:
                lookup[(str(centro).strip(), str(fab).strip())] = float(valor)
    return lookup


def procesar(tipo_id: str, file_storage) -> pd.DataFrame:
    """
    Punto de entrada principal: recibe el Excel del usuario (file-like) y
    devuelve el DataFrame ya ampliado, listo para exportar.
    """
    cfg = cargar_config(tipo_id)
    input_df = _leer_input(file_storage, cfg)
    ref_df_completa = _cargar_referencia(cfg["reference_file"], cfg.get("reference_sheet"))

    key_input = cfg["key_input"]
    key_ref = cfg["key_reference"]
    fallback = cfg.get("fallback_value")
    cols_from_ref = cfg["output_columns_from_reference"]
    text_columns = cfg.get("text_columns", [])
    filial_column = cfg.get("filial_column")
    reference_filter = cfg.get("reference_filter")

    if reference_filter:
        ref_df_completa = ref_df_completa[
            ref_df_completa[reference_filter["column"]] == reference_filter["value"]
        ]

    disponibilidad = _cargar_disponibilidad(cfg)
    bloqueado_valor = cfg.get("disponibilidad", {}).get("valor_bloqueado", 0)

    corr_cfg = cfg.get("correlativo", {"enabled": False})

    filas_salida = []
    avisos = []

    def _filtrar_por_disponibilidad(matches_df, fabricante):
        """Saca del match los centros que la matriz de disponibilidad marca
        como bloqueados para ese fabricante, aunque la tabla de referencia
        traiga una fila (dato posiblemente mal cargado ahí)."""
        if disponibilidad is None or matches_df.empty:
            return matches_df, []
        centros_excluidos = []
        indices_ok = []
        for idx, fila_ref in matches_df.iterrows():
            centro = str(fila_ref["CENTRO"]).strip()
            estado = disponibilidad.get((centro, str(fabricante).strip()))
            if estado == bloqueado_valor:
                centros_excluidos.append(centro)
            else:
                indices_ok.append(idx)
        return matches_df.loc[indices_ok], centros_excluidos

    def _fabricante_bloqueado_en_toda_la_filial(ref_df_filial, fabricante) -> bool:
        """
        True si la matriz de disponibilidad conoce a este fabricante para AL
        MENOS uno de los centros de esta filial, y lo marca en 0 para TODOS
        ellos (nunca en 1 o 0.5). Cubre el caso en que la tabla de referencia
        fabricante->centro/cebe ni siquiera tiene una fila para esta
        combinación (por eso no basta con filtrar matches_crudos): un
        fabricante conocido-pero-no-habilitado en esta filial no debe caer en
        el comodín F9999, tiene que rechazarse.
        Si la matriz no conoce el fabricante para ningún centro de la filial,
        devuelve False (es un fabricante nuevo/no catalogado: sigue el
        comportamiento normal de fallback).
        """
        if disponibilidad is None or ref_df_filial.empty:
            return False
        centros_filial = ref_df_filial["CENTRO"].astype(str).str.strip().unique()
        estados = [
            disponibilidad[(centro, str(fabricante).strip())]
            for centro in centros_filial
            if (centro, str(fabricante).strip()) in disponibilidad
        ]
        if not estados:
            return False
        return all(e == bloqueado_valor for e in estados)

    for _, fila_input in input_df.iterrows():
        valor_key = fila_input[key_input]

        ref_df = ref_df_completa
        if filial_column:
            filial_fila = str(fila_input.get("FILIAL CODIGO", "")).strip()
            ref_df = ref_df[ref_df[filial_column] == filial_fila]

        matches_crudos = ref_df[ref_df[key_ref] == valor_key]
        matches, centros_excluidos = _filtrar_por_disponibilidad(matches_crudos, valor_key)

        if not matches_crudos.empty and matches.empty:
            # El fabricante SÍ tiene fila(s) en la tabla de referencia, pero la
            # matriz de disponibilidad los bloquea a todos: no caer al comodín,
            # sería disfrazar un fabricante conocido-pero-no-habilitado como
            # "todas/otras marcas".
            avisos.append(
                f"Fabricante '{valor_key}' no habilitado para {', '.join(sorted(set(centros_excluidos)))} "
                f"(matriz de disponibilidad) — fila omitida."
            )
            continue

        if matches.empty and _fabricante_bloqueado_en_toda_la_filial(ref_df, valor_key):
            avisos.append(
                f"Fabricante '{valor_key}' no habilitado en ningún centro de "
                f"'{filial_fila if filial_column else ''}' (matriz de disponibilidad) — fila omitida."
            )
            continue

        uso_fallback = False
        if matches_crudos.empty and fallback:
            matches_crudos = ref_df[ref_df[key_ref] == fallback]
            matches, _ = _filtrar_por_disponibilidad(matches_crudos, fallback)
            uso_fallback = True

        if matches.empty:
            avisos.append(f"Fabricante '{valor_key}' no encontrado — fila omitida.")
            continue

        if centros_excluidos:
            avisos.append(
                f"Fabricante '{valor_key}': se excluyeron {', '.join(sorted(set(centros_excluidos)))} "
                f"(no habilitados), se amplió al resto."
            )

        # Un mismo material conserva UN solo número, repetido en todas sus
        # filas ampliadas (no uno distinto por centro).
        numero_material = None
        if corr_cfg.get("enabled"):
            numero_material = correlativo.siguiente_numero(
                nombre_contador=corr_cfg["nombre"],
                range_min=corr_cfg["range_min"],
                range_max=corr_cfg["range_max"],
                tipo=tipo_id,
                texto_breve=str(fila_input.get("TEXTO BREVE", "")),
                fabricante_codigo=str(valor_key),
            )

        for _, fila_ref in matches.iterrows():
            fila_out = {}
            # Copiar valores preservando tipos de texto
            for col in fila_input.index:
                if col in text_columns:
                    fila_out[col] = str(fila_input[col])
                else:
                    fila_out[col] = fila_input[col]
            
            for col in cols_from_ref:
                if col in text_columns:
                    fila_out[col] = str(fila_ref.get(col, ""))
                else:
                    fila_out[col] = fila_ref.get(col)
            
            if numero_material is not None:
                fila_out[corr_cfg["column_name"]] = numero_material
            if uso_fallback:
                fila_out["_FALLBACK_USADO"] = "SI"
            filas_salida.append(fila_out)

    if not filas_salida:
        raise InputInvalidoError(
            "No se generó ninguna fila de salida. " + " ".join(avisos)
        )

    resultado = pd.DataFrame(filas_salida)
    
    # Aplicar tipos de datos específicos
    for col in text_columns:
        if col in resultado.columns:
            resultado[col] = resultado[col].astype(str)
    
    resultado.attrs["avisos"] = avisos
    return resultado
